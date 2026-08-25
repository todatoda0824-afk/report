# -*- coding: utf-8 -*-
"""
경쟁사 분기 실적/공시 자동 집계
================================
config.py 의 설정을 읽어 DART Open API에서
  1) 연결기준 매출액·영업이익
  2) 최근 주요 공시현황
  3) 법인별(별도기준) 매출액·당기순이익
을 조회해 엑셀 파일 하나로 만듭니다.

실행: run.bat (윈도우) 더블클릭, 또는 터미널에서 `python generate_report.py`
"""

import sys
import os
import datetime
import traceback

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import config
except (SyntaxError, UnicodeDecodeError):
    print("=" * 60)
    print("config.py 파일을 읽는 중 오류가 발생했습니다.")
    print("메모장으로 config.py를 열어 '다른 이름으로 저장' 시")
    print("인코딩을 'UTF-8'로 선택해서 다시 저장한 뒤 실행해주세요.")
    print("=" * 60)
    input("엔터를 누르면 종료합니다...")
    sys.exit(1)

try:
    import OpenDartReader
except ImportError:
    try:
        # 최신 배포판은 패키지 폴더명이 소문자 opendartreader 로 바뀐 경우가 있음
        from opendartreader import OpenDartReader
    except ImportError:
        try:
            import opendartreader as _odr_module
            OpenDartReader = getattr(_odr_module, "OpenDartReader", None)
            if OpenDartReader is None:
                raise ImportError
        except ImportError:
            print("OpenDartReader가 설치되어 있지 않습니다. 아래 명령으로 설치 후 다시 실행하세요:")
            print("    pip install OpenDartReader")
            sys.exit(1)


# ------------------------------------------------------------------
# 분기 <-> DART 보고서코드 매핑
# ------------------------------------------------------------------
REPRT_CODE = {
    1: "11013",  # 1분기보고서
    2: "11012",  # 반기보고서
    3: "11014",  # 3분기보고서
    4: "11011",  # 사업보고서 (연간)
}
QUARTER_LABEL = {1: "1분기", 2: "반기(2분기)", 3: "3분기", 4: "사업보고서(연간)"}


def get_latest_available_quarter(today=None):
    """
    법정 제출기한(분기·반기 45일, 사업보고서 90일)에 여유(버퍼)를 두고
    '오늘 기준으로 이미 공시되었을 가능성이 높은' 최신 분기를 추정합니다.
    회사마다 실제 제출일은 다를 수 있으니, 정확한 분기를 원하면
    config.py 에서 AUTO=False 로 두고 직접 지정하세요.
    """
    if today is None:
        today = datetime.date.today()
    y = today.year

    candidates = [
        (y - 1, 4, datetime.date(y, 4, 10)),        # 전년도 사업보고서
        (y, 1, datetime.date(y, 5, 20)),             # 올해 1분기
        (y, 2, datetime.date(y, 8, 20)),             # 올해 반기
        (y, 3, datetime.date(y, 11, 20)),            # 올해 3분기
        (y, 4, datetime.date(y + 1, 4, 10)),          # 올해 사업보고서(내년 발표)
    ]
    valid = [(yy, q) for yy, q, due in candidates if due <= today]
    if not valid:
        return (y - 2, 4)
    return max(valid, key=lambda x: (x[0], x[1]))


def safe_amount(val):
    """'1,234,567' 같은 문자열을 정수로. 실패하면 None."""
    if val is None:
        return None
    try:
        s = str(val).replace(",", "").strip()
        if s in ("", "-"):
            return None
        return int(s)
    except (ValueError, TypeError):
        return None


def to_eok(won):
    """원 -> 억원 (소수 1자리 반올림)"""
    if won is None:
        return None
    return round(won / 100_000_000, 1)


def find_account(df, keywords):
    """account_nm 컬럼에서 keywords 중 하나라도 포함하는 첫 행을 반환."""
    if df is None or len(df) == 0:
        return None
    for kw in keywords:
        hit = df[df["account_nm"].astype(str).str.replace(" ", "").str.contains(kw, na=False)]
        if len(hit) > 0:
            return hit.iloc[0]
    return None


def fetch_corp_info(dart, identifier):
    """dart.company()로 실제 매칭된 법인 정보(종목코드, 정식명칭 등)를 가져옵니다.
    이름이 같은 다른 법인이 잘못 매칭되지 않았는지 검증하는 용도입니다."""
    try:
        info = dart.company(identifier)
    except Exception:
        return None
    return info


def extract_field(info, key):
    if info is None:
        return ""
    try:
        if hasattr(info, "get"):
            val = info.get(key, "")
            return val if val else ""
    except Exception:
        pass
    try:
        return info[key]
    except Exception:
        return ""


def get_identifier(corp_name):
    """config.py의 COMPANY_CODE_OVERRIDE에 종목코드가 지정된 회사는 코드로,
    아니면 회사명 그대로 조회에 사용합니다. 종목코드를 쓰면 동명의 다른 법인과
    섞일 위험이 없어집니다."""
    override = getattr(config, "COMPANY_CODE_OVERRIDE", {}) or {}
    return override.get(corp_name, corp_name)


def get_periods():
    """조회할 (연도, 분기) 목록을 반환합니다."""
    if getattr(config, "MULTI_PERIOD", False):
        return list(config.PERIODS)
    if config.AUTO:
        return [get_latest_available_quarter()]
    return [(config.MANUAL_YEAR, config.MANUAL_QUARTER)]


def fetch_financials(dart, identifier, year, reprt_code, fs_div):
    """fs_div: 'CFS'(연결) 또는 'OFS'(별도). 실패 시 None 반환."""
    try:
        df = dart.finstate(identifier, year, reprt_code=reprt_code, fs_div=fs_div)
    except Exception as e:
        if config.DEBUG:
            print(f"    [진단] finstate() 실패 ({identifier}, {fs_div}, {year}): {type(e).__name__}: {e}")
        try:
            df = dart.finstate_all(identifier, year, reprt_code=reprt_code, fs_div=fs_div)
            if config.DEBUG:
                print(f"    [진단] finstate_all()로 재시도 성공 ({identifier}, {fs_div})")
        except Exception as e2:
            if config.DEBUG:
                print(f"    [진단] finstate_all()도 실패 ({identifier}, {fs_div}): {type(e2).__name__}: {e2}")
            return None
    if df is None or len(df) == 0:
        return None
    return df


def fetch_disclosures(dart, identifier, start_date, end_date, max_rows):
    try:
        df = dart.list(identifier, start=start_date, end=end_date)
    except Exception:
        return None
    if df is None or len(df) == 0:
        return None
    df = df.sort_values("rcept_dt", ascending=False).head(max_rows)
    return df


# ------------------------------------------------------------------
# 엑셀 스타일 헬퍼
# ------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
FAIL_FONT = Font(name="Arial", size=10, color="C00000")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_sheet(wb, title, headers, rows, col_widths=None, money_cols=None):
    ws = wb.create_sheet(title)
    money_cols = money_cols or set()

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.font = FAIL_FONT if (isinstance(val, str) and val.startswith("조회실패")) else BODY_FONT
            if c in money_cols and isinstance(val, (int, float)):
                cell.number_format = "#,##0.0"
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    if col_widths:
        for c, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
    return ws


def main():
    if not config.DART_API_KEY or "여기에" in config.DART_API_KEY:
        print("=" * 60)
        print("config.py 에 DART_API_KEY 를 먼저 입력해주세요.")
        print("https://opendart.fss.or.kr 에서 무료로 발급받을 수 있습니다.")
        print("=" * 60)
        input("엔터를 누르면 종료합니다...")
        sys.exit(1)

    dart = OpenDartReader(config.DART_API_KEY)

    periods = get_periods()
    multi = getattr(config, "MULTI_PERIOD", False)

    print(f"조회 대상 분기: {', '.join(f'{y}년 {QUARTER_LABEL[q]}' for y, q in periods)}")
    print(f"조회 회사 수: {len(config.COMPANY_LIST)}개")
    print("-" * 60)

    today = datetime.date.today()
    if multi:
        earliest_year = min(y for y, _ in periods)
        start_date = datetime.date(earliest_year, 1, 1).strftime("%Y%m%d")
    else:
        start_date = (today - datetime.timedelta(days=config.DISCLOSURE_LOOKBACK_DAYS)).strftime("%Y%m%d")
    end_date = today.strftime("%Y%m%d")

    consolidated_rows = []
    separate_rows = []
    disclosure_rows = []
    failed_companies = []

    total = len(config.COMPANY_LIST)
    for i, corp_name in enumerate(config.COMPANY_LIST, start=1):
        identifier = get_identifier(corp_name)
        print(f"[{i}/{total}] {corp_name} ({identifier}) 조회 중...")

        # 매칭 법인 검증 (종목코드/정식명칭) --------------------------------
        info = fetch_corp_info(dart, identifier)
        stock_code = extract_field(info, "stock_code") or (identifier if identifier != corp_name else "")
        dart_matched_name = extract_field(info, "corp_name") or ""

        for year, quarter in periods:
            reprt_code = REPRT_CODE[quarter]

            # 1) 연결기준 매출액/영업이익 ------------------------------------
            cfs = fetch_financials(dart, identifier, year, reprt_code, "CFS")
            rev = find_account(cfs, ["매출액", "영업수익"]) if cfs is not None else None
            opi = find_account(cfs, ["영업이익"]) if cfs is not None else None

            if rev is None and opi is None:
                consolidated_rows.append([
                    i, corp_name, stock_code, dart_matched_name, year, QUARTER_LABEL[quarter],
                    "조회실패(연결)", "조회실패(연결)", None, None,
                ])
                failed_companies.append(f"{corp_name} {year}년{QUARTER_LABEL[quarter]} (연결재무제표)")
            else:
                rev_this = to_eok(safe_amount(rev["thstrm_amount"])) if rev is not None else None
                rev_prev = to_eok(safe_amount(rev["frmtrm_amount"])) if rev is not None else None
                opi_this = to_eok(safe_amount(opi["thstrm_amount"])) if opi is not None else None
                opi_prev = to_eok(safe_amount(opi["frmtrm_amount"])) if opi is not None else None
                consolidated_rows.append([
                    i, corp_name, stock_code, dart_matched_name, year, QUARTER_LABEL[quarter],
                    rev_this, opi_this, rev_prev, opi_prev,
                ])

            # 2) 법인별(별도) 매출액/당기순이익 -------------------------------
            ofs = fetch_financials(dart, identifier, year, reprt_code, "OFS")
            rev_o = find_account(ofs, ["매출액", "영업수익"]) if ofs is not None else None
            ni_o = find_account(ofs, ["당기순이익", "반기순이익", "분기순이익"]) if ofs is not None else None

            if rev_o is None and ni_o is None:
                separate_rows.append([
                    i, corp_name, stock_code, dart_matched_name, year, QUARTER_LABEL[quarter],
                    "조회실패(별도)", "조회실패(별도)",
                ])
                failed_companies.append(f"{corp_name} {year}년{QUARTER_LABEL[quarter]} (별도재무제표)")
            else:
                separate_rows.append([
                    i, corp_name, stock_code, dart_matched_name, year, QUARTER_LABEL[quarter],
                    to_eok(safe_amount(rev_o["thstrm_amount"])) if rev_o is not None else None,
                    to_eok(safe_amount(ni_o["thstrm_amount"])) if ni_o is not None else None,
                ])

        # 3) 최근 공시현황 (분기 루프 밖에서 한 번만) -------------------------
        disc = fetch_disclosures(dart, identifier, start_date, end_date, config.DISCLOSURE_MAX_PER_COMPANY)
        if disc is None or len(disc) == 0:
            disclosure_rows.append([corp_name, stock_code, "조회실패 또는 최근 공시 없음", "", "", ""])
        else:
            for _, d in disc.iterrows():
                rcept_no = d.get("rcept_no", "")
                link = f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rcept_no}" if rcept_no else ""
                disclosure_rows.append([
                    corp_name,
                    stock_code,
                    d.get("rcept_dt", ""),
                    d.get("report_nm", ""),
                    d.get("flr_nm", ""),
                    link,
                ])

    # ------------------------------------------------------------------
    # 엑셀 작성
    # ------------------------------------------------------------------
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    if multi:
        period_label = f"{periods[0][0]}Q{periods[0][1]}-{periods[-1][0]}Q{periods[-1][1]}"
    else:
        y0, q0 = periods[0]
        period_label = f"{y0}_{QUARTER_LABEL[q0].replace('(', '').replace(')', '')}"
    out_path = os.path.join(
        config.OUTPUT_DIR,
        f"경쟁사현황_{period_label}_{today.strftime('%Y%m%d')}.xlsx",
    )

    wb = Workbook()
    wb.remove(wb.active)

    write_sheet(
        wb, "연결_매출영업이익",
        ["No", "회사명", "종목코드", "DART매칭명", "연도", "분기", "매출액(억원,당기)", "영업이익(억원,당기)",
         "매출액(억원,전년동기)", "영업이익(억원,전년동기)"],
        consolidated_rows,
        col_widths=[5, 16, 10, 16, 8, 12, 16, 16, 18, 18],
        money_cols={7, 8, 9, 10},
    )

    write_sheet(
        wb, "법인별_매출당기순이익",
        ["No", "회사명", "종목코드", "DART매칭명", "연도", "분기", "매출액(억원,별도)", "당기순이익(억원,별도)"],
        separate_rows,
        col_widths=[5, 16, 10, 16, 8, 12, 16, 18],
        money_cols={7, 8},
    )

    write_sheet(
        wb, "주요공시현황",
        ["회사명", "종목코드", "접수일자", "보고서명", "공시대상회사", "원문링크"],
        disclosure_rows,
        col_widths=[16, 10, 12, 45, 16, 55],
    )

    # 요약/메타 시트
    meta_rows = [
        ["생성일시", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["조회 기준", ", ".join(f"{y}년 {QUARTER_LABEL[q]}" for y, q in periods)],
        ["조회 회사 수", len(config.COMPANY_LIST)],
        ["조회 실패 항목 수", len(failed_companies)],
    ]
    write_sheet(wb, "요약", ["항목", "값"], meta_rows, col_widths=[20, 40])
    if failed_companies:
        write_sheet(
            wb, "조회실패목록",
            ["실패 항목"],
            [[f] for f in failed_companies],
            col_widths=[45],
        )

    wb.save(out_path)

    print("-" * 60)
    print(f"완료: {out_path}")
    if failed_companies:
        print(f"※ {len(failed_companies)}건 조회 실패 - '조회실패목록' 시트 참고")
    print("-" * 60)


if __name__ == "__main__":
    try:
        main()
    except Exception:
        print("오류가 발생했습니다:")
        traceback.print_exc()
    finally:
        input("엔터를 누르면 창을 닫습니다...")
